import os
import re
import math
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font

def remove_paragraph_indents(text):
    cleaned_lines = [line.lstrip() for line in text.splitlines()]
    return '\n'.join(cleaned_lines)

# Список стоп-слов
stop_words = {
    'کجا', 'نبودی', 'بودی', 'شدی', 'nn', 'آنجا', 'آنکه', 'های', 'ست', 'کرده',
    'ازین', 'ازان', 'باز', 'گردد', '،', 'گشت', 'فی', 'زین', 'کو', 'هست', 'چنین',
    'برای', 'شود', 'کس', 'کز', 'بهر', 'همچو', 'سوی', 'پی', 'همی', 'شمارهٔ', 'یا',
    'چو', 'مر', 'گفت', 'برات', 'چون', 'هیچ', 'صد', 'دارد', 'داری', 'نیست', 'است',
    'گوید', 'چی', 'چه', 'اگر', 'هر', 'را', 'از', 'ز', 'چهار', 'سه', 'یک', 'شما',
    'وی', 'انها', 'ما', 'او', 'تو', 'من', 'دو', 'پیش', 'گر', 'می', 'و', 'در', 'کی',
    'شد', 'کرد', 'گرچه', 'گاه', 'گه', 'زان', 'که', 'به', 'بر', 'آن', 'تا', 'این',
    'اندر', 'آندر', 'اینجا', 'درین', 'بود', 'بی', 'آمد', 'چون', 'با', 'ای', 'هم',
    'باشد', 'ترا', '-', ':', 'سر', 'شماره', 'شده', 'شد', 'ام', 'وز'
}

# Ключевые слова для анализа
keywords = ['کلام','کلمه','امر','جبر','اختیار', 'قادر', 'قضا', 'علم', 'عمل', 'ارادت', 'قدرت']

# Путь к корпусу
folder_path = r'C:\Users\Lenovo\Documents\корпус персидской поэзии\attar'
texts = []

def clean_text(text, keep_words=None):
    """Очищает текст, сохраняя ключевые слова"""
    if keep_words is None:
        keep_words = set()
    
    # Исправленные регулярные выражения (добавлены r перед строками)
    tokens = re.split(r'(\s+)', text)  # Исправлено: \s вместо s
    cleaned_tokens = []
    
    for token in tokens:
        if token.strip() in keep_words:
            cleaned_tokens.append(token)
        elif token.isspace():
            cleaned_tokens.append(token)
        else:
            cleaned_token = re.sub(r'\d+', '', token)  # Исправлено: \d вместо d
            cleaned_token = re.sub(r'[^\w\s]', '', cleaned_token)  # Исправлено: \w\s вместо ws
            cleaned_tokens.append(cleaned_token)
    
    return ''.join(cleaned_tokens)

# Загрузка и очистка текстов
for filename in os.listdir(folder_path):
    if filename.endswith('.txt'):
        with open(os.path.join(folder_path, filename), 'r', encoding='utf-8') as file:
            content = file.read()
            content = clean_text(content, set(keywords))
            cleaned_content = remove_paragraph_indents(content)
            texts.append(cleaned_content)

# Фильтрация слов
def filter_words(text, stop_words, keywords):
    words = text.split()
    return [word for word in words if word in keywords or (word not in stop_words and word.strip() != '')]

corpus_words = []
for text in texts:
    corpus_words.extend(filter_words(text, stop_words, keywords))

# Подсчет частот
word_freq = defaultdict(int)
for word in corpus_words:
    word_freq[word] += 1

total_words = len(corpus_words)

def calculate_dpf(corpus_words, keywords, window_size=10):
    cooccurrence = defaultdict(lambda: defaultdict(int))
    dpf_results = defaultdict(dict)

    # 1. Подсчет совместной встречаемости
    for i, word in enumerate(corpus_words):
        if word in keywords:
            start = max(0, i - window_size)
            end = min(len(corpus_words), i + window_size + 1)
            
            for j in range(start, end):
                if j != i and corpus_words[j] in keywords:
                    cooccurrence[word][corpus_words[j]] += 1

    # 2. Вычисление DPF с улучшенной формулой
    for word1 in keywords:
        for word2 in keywords:
            if word1 != word2:
                # Количество совместных вхождений
                cooc_count = cooccurrence[word1][word2]
                
                # Частоты слов
                freq1 = word_freq[word1]
                freq2 = word_freq[word2]
                
                if freq1 > 0 and freq2 > 0 and total_words > 0:
                    # Вероятности
                    p_word2_given_word1 = cooc_count / freq1
                    p_word2 = freq2 / total_words
                    
                    # Расчет DPF
                    if p_word2 > 0:
                        dpf = math.log2((p_word2_given_word1 + 0.0001) / (p_word2 + 0.0001))  # Добавлен сглаживающий фактор
                        direction = "←" if dpf < 0 else "→"
                        dpf_results[word1][word2] = (abs(dpf), direction)
                    else:
                        dpf_results[word1][word2] = (0, "→")
                else:
                    dpf_results[word1][word2] = (0, "→")

    return dpf_results

# Вычисляем DPF
dpf_results = calculate_dpf(corpus_words, keywords)

# Создание Excel-файла с результатами
def create_dpf_excel(results, filename="dpf_results.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "DPF Results"

    # Заголовки
    ws.cell(row=1, column=1, value="Keyword")
    for col, keyword in enumerate(keywords, start=2):
        ws.cell(row=1, column=col, value=keyword).font = Font(bold=True)

    # Данные
    for row, word1 in enumerate(keywords, start=2):
        ws.cell(row=row, column=1, value=word1).font = Font(bold=True)
        for col, word2 in enumerate(keywords, start=2):
            if word1 == word2:
                ws.cell(row=row, column=col, value="-")
            else:
                dpf, direction = results.get(word1, {}).get(word2, (0, ""))
                ws.cell(row=row, column=col, value=f"{direction}{dpf:.2f}")

    # Автоподбор ширины столбцов
    for col in range(1, len(keywords) + 2):
        ws.column_dimensions[chr(64 + col)].width = 12

    wb.save(filename)
    return filename

# Сохранение результатов
excel_file = create_dpf_excel(dpf_results)
print(f"DPF результаты сохранены в файл: {excel_file}")

print("\nDPF матрица (направление → означает word1 → word2):")
print("".ljust(12), end="")
for word in keywords:
    print(word[:10].ljust(12), end="")
print()

for word1 in keywords:
    print(word1[:10].ljust(12), end="")
    for word2 in keywords:
        if word1 == word2:
            print("-".ljust(12), end="")
        else:
            dpf, direction = dpf_results.get(word1, {}).get(word2, (0, ""))
            print(f"{direction}{dpf:.2f}".ljust(12), end="")
    print()